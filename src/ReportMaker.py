from pathlib import Path
from tkinter import messagebox

import pandas as pd

from .Functions import time_responser

class DataFeeder:
    """
    Creates two dataframes with the evidence data from the "Passed" and "Defects" folders in the Worktree. 
    This class works together with the 'ReportMaker' class.
    """

    def __init__(self, prefix:str, path:Path):
        self.prefix = prefix
        self.path = path
    
    @classmethod
    def passed_headers(cls):
        headers = ["Slot", "OS", "Clone","Test","Retest"]
        return headers

    @classmethod
    def defect_headers(cls):
        headers = ["Slot", "OS", "Clone","Test","Defect ID"]
        return headers
    
    def create_row(self,path:Path):
        """
        Creates an array of data based on the files in the directory passed to the method.
        """
        import re
        try:
            final_path = path/self.prefix 
            data = [f for f in final_path.iterdir() if f.is_file() and f.suffix.lower() in {'.mp4','.mov','.zip'}]
            rows =[]
            for f in data:
                parts = re.split(r"-+",f.stem)
        
                if self.prefix == "Passed":
                    if len(parts) == 4:
                        parts.append(None) 
                    elif len(parts) != 5:
                        print(f"Skipping file (unexpected format): {f.name}")
                        continue

            # If it's "Defects", all 5 parts are mandatory
                elif self.prefix == "Defects" and len(parts) != 5:
                    print(f"Skipping file (invalid defect format): {f.name}")
                    continue

                rows.append(parts)
            
            print(rows)
            return rows
        except ValueError:
            pass
    
    def to_dataframe(self, path:Path ,headers: list[str])-> pd.DataFrame:
        """
        Passes the data into a dataframe
        """
        try:
            rows = self.create_row(path)
            print("to dataframe ok")
            return pd.DataFrame(rows, columns=headers)

        except Exception as e:
            print(e)


class ReportMaker:
    """
    Creates an .xlsx file, printing two Pandas Dataframes on two different sheets. 
    The data in the dataframes must be supplied via the CSV_File class.
    """

    def __init__(self,path: Path):
        self.date_str = time_responser('date')
        self.path = path
        self.filename = self.path/f'Report_{self.date_str}.xlsx'


    def data_feed(self, df_passed:pd.DataFrame, df_defect:pd.DataFrame):
        """
        Creates the .xlsx file and loads the data from the CSV into it.
        """
        import openpyxl as op
        try:
            
            # CReates the xlsx file and writes the Passed data 
            df_passed.to_excel(self.filename, index=False, sheet_name="Passed")
            with pd.ExcelWriter(self.filename, engine="openpyxl", mode="a") as writer:

                 # Writes the Defects data
                df_defect.to_excel(writer, index=False, sheet_name="Defects")
            
            # Loads the sheets to an Openpyxl workbook
            wb = op.load_workbook(self.filename)

            for sheet_name in ["Passed","Defects"]:
            # Defines the two sheets inside the file and formats the columns
                ws = wb[sheet_name]
                self.adjust_column_width(ws)
                self.convert_to_table(ws)

            # Saves the file
            wb.save(self.filename)
            messagebox.showinfo(title="Report creato!", message=f"Il report {self.date_str} è stato creato con successo!")

        except pd.errors.ParserError:
            messagebox.showerror(title="Error!", message="A file does not respect the format!")

    def adjust_column_width(self, worksheet):
        """
        Adapts the width of the columns based on the content.
        """
        for col in worksheet.columns:
            try:
                max_length = 0
                column = col[0].column_letter 
                for cell in col:
                        if len(cell.value) > max_length:
                            max_length = max(max_length,len(cell.value))
                adjusted_width = (max_length + 2) 
                worksheet.column_dimensions[column].width = adjusted_width
                # It's gonna raise a TypeError exception anytime there's an empty field in the CSV
                # That's gonna happen sometimes, so it's pass.
            except TypeError:
                pass
    
    def convert_to_table(self, worksheet):
        """
        Converts the used range of the worksheet into an Excel table with styling.
        """
        from openpyxl.worksheet.table import Table, TableStyleInfo
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if max_row < 2:
            return  # No data to table-ify

        ref = f"A1:{worksheet.cell(row=max_row, column=max_col).coordinate}"
        table = Table(displayName=f"Table_{worksheet.title}", ref=ref)

        # Apply a table style
        style = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        table.tableStyleInfo = style

        worksheet.add_table(table)